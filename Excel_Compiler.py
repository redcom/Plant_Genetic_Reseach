#!/usr/bin/env python

import os
import fnmatch
import numpy as np
import itertools
from openpyxl import Workbook

top_folder = raw_input("Enter top folder: ")

print("***\n**\n*\nCompiling Stat files in " + top_folder + "\n.\n.")

match_list = []
stat_file = top_folder + "\Complete_Stats.xlsx"

# Find all files with stat data
for root, dirnames, filenames in os.walk(top_folder):
    for filename in fnmatch.filter(filenames, 'Stats_*.csv'):
        match_list.append(os.path.join(root, filename))

all_stats = {}

plot_amount = 0

# find all of the plots avaialble. since some files are missing a few plots
# if the file with the most amount of plots is missing a few, then trouble
for f in range(len(match_list)):

    data = np.genfromtxt(match_list[f], dtype=None, delimiter=",",
                         names=True, skip_header=0)

    print match_list[f]
    if len(data[data.dtype.names[0]]) > plot_amount:

        plot_ids = tuple(data[data.dtype.names[0]])
        all_stats.update({"PLOTS": plot_ids})
        plot_amount = len(plot_ids)

# Throw all data into a big dictionary where trait_sensor_run is the key
for g in range(len(match_list)):

    data = np.genfromtxt(match_list[g], dtype=None, delimiter=",",
                         names=True, skip_header=0)

    plot_check = {}
    trait_avgs = [np.nan] * len(plot_ids)

    for row in range(len(data)):

        plot_check.update({data[row][0]: data[row][1]})

    for h in range(len(plot_ids)):

        pid = plot_ids[h]

        if pid in plot_check.keys():

            trait_avgs[h] = plot_check[pid]

    all_stats.update({data.dtype.names[1]: trait_avgs})

key_headers = all_stats.keys()
key_headers.sort()
key_headers.remove("PLOTS")

key_headers_split = []
group_keys = []

book = Workbook()
ws = book.active

# Remove default sheet
bad_sheet = book.get_sheet_by_name('Sheet')
book.remove_sheet(bad_sheet)

# create list for which to group sheets by [[CIRE, CC], [Target, Cel, IRT]...]
for k in range(len(key_headers)):
    key_headers_split.append(key_headers[k].split("_"))

# group headers by name (first two elements in key_headers_split)
for key, group in itertools.groupby(key_headers_split, lambda x: x[0:2]):

    sheet = book.create_sheet(title="_".join(key))
    all_runs = ["PLOTS"]

    for column_name in group:

        all_runs.append("_".join(column_name))

    for r in range(len(plot_ids)):

        combined = []

        for col in range(len(all_runs)):

            # First column is always plot ids, others are the runs
            if col == 0:
                combined.append(plot_ids)
            else:
                combined.append(all_stats[all_runs[col]])

        # insert header first
        if r == 0:

            for h in range(len(all_runs)):
                sheet.cell(column=h+1, row=r+1, value=str(all_runs[h]))

        # Insert data into each cell
        for c in range(len(combined)):
            sheet.cell(column=c+1, row=r+2, value=combined[c][r])

book.save(filename=stat_file)

print "Spreadsheet created!\n*\n**\n***"
