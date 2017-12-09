#!/usr/bin/env python
import sys
import os
import numpy as np
import csv
import tkinter
import warnings
import pandas
import sqlite3
import fnmatch
import time
import datetime
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from multiprocessing import Pool
import itertools
from openpyxl import Workbook


class User_Interface(object):
    """This class is a TKinter-based GUI.

    utilizing grid manager, rather than pack() for now
    """

    def __init__(self):
        """This will execute if User_Interface() is called

        Will prompt user to enter Project Name, Workspace,
        Plant Data Shapefile, and Image File
        """

        # creates a master window needed for the widgets to exist on
        self.root = tkinter.Tk()
        self.root.wm_title("Files for The_Cleaners")

        info_row = 0
        info_column = 0

        info_text = "Instruction on how to\nuse this GUI/script"
        self.label_info = tkinter.Label(self.root, text=info_text)
        # grid() assigns locations. sticky moves it NSE or W within cell
        self.label_info.grid(row=info_row, column=info_column,
                             sticky=tkinter.E)

        explanation = """
*******************************************************************************
        This area is provided to assist the user on how to execute this GUI.
*******************************************************************************
Scroll down for more...

This software is to perform cleaning and analysis on the data.
The files that can be operated on are:
    CC_plots.csv
    IRT_forward_plots.csv
    IRT_nadir_plots.csv
    Honey_plots.csv
    Pep_plots.csv

GUI explanation:
The first entry box is for the directory in which the script is to operate in.
It can house a singular run or multiple runs.
The requirement is for "_Run##_" to appear anywhere within the path name.
For example: "E:\htp\F120\_QC1 Output\2017_01_04_F120_Run01_DOY004"

In the next row, an option for "Stressors" refers to the first letter character
    in the plot name.
    For example, in the plot "W325N", you would enter "W".
    It is acceptable to have more than one stressor. When listing the stressors
    be sure not to include any spaces between characters.
    For example "W,D"
The last option in the same row is for the different species of plants.
    It is also acceptable to have more than one species.
    For example, for plot "W325N", the entry would be "N"
    Or if multiple, "N,J,C"

The next segment is to toggle which sensor to run, your 4 choices are:
    Crop Circle (Toggle CC_plot), Infrared Temperature (Toggle IRT_plot),
    Honeywell (Toggle Honey_plot), and Pepperl (Toggle Pep_plot)

    Simply click the gray bar to toggle which sensor you'd like to analyze.
    When you've clicked it, it should highlight green to let you know that
    analysis for that sensor will be executed.
    Note: The files must exist and have the same name as the files listed above

Within each section is the option to:
    -Define "Range". Here you may enter any two numbers that you'd like to
        establish as the range. The form is "0,1" without spaces.
        Note: The range given here defines the range of the Y-axis on the plots
        -- Don't Define Range.  A check box in the same row to let the script
           know you would not like to define any range
    -Plot Data. An option to create boxplots, per plot, of the sensor data.
        Note: Selecting this will also increase the amount of time
        the script takes to run. (Approx +1 min per species with 350 plots)
        -- If Toggle Range SD is selected, sensors with a mean outside of 1 SD
           of the plot mean will not show up because they have been thrown out.
           Also, a red dashed line and pink line will be drawn representing
           +/- 1 standard deviation and the mean, respectively.
    -Toggle Range SD. Refering to an analysis of taking the Standard Deviation
          of a plot's sensors, finding the mean, taking a range defined as the
          mean +\- 1 Standard Deviation, then taking the mean of the points
          found within that range and calculating their standard deviation.
          If the mean value of each sesnor is outside 1 standard deviation of
          the new mean, then that sensor will be removed.
    -Remove Zeros. This will remove any values that are exactly zero
    -Reanalyze Stat. This option is for if you have already run the script
         and have visually identified bad sensors that need to be removed. To
         effectively remove the sensors, there will be a file within
         \Cleaned_Data\ called (Sensor)_preened.csv, open it and enter the
         sensors into the file, abiding by the format laid out.


To execute this file in Command Prompt, open Command Prompt:

$ cd to/the/script
$ python The_Cleaners.py
        """

        self.textbox = tkinter.Text(
                     self.root, height=5,
                     width=125, wrap=tkinter.WORD)
        self.textbox.grid(row=info_row, column=info_column+1, columnspan=5)

        self.scrollbar = tkinter.Scrollbar(self.root)
        self.scrollbar.grid(row=info_row, column=info_column+6,
                            sticky=tkinter.N + tkinter.S + tkinter.W)

        self.textbox.config(yscrollcommand=self.scrollbar.set)
        self.scrollbar.config(command=self.textbox.yview)

        self.textbox.insert(tkinter.END, explanation)

        filler_text = "            "
        self.label_run = tkinter.Label(self.root, text=filler_text)
        self.label_run.grid(row=1, column=0)

        fd_row = 2
        fd_column = 0
        # Creates widgets for a label and space to input a project name
        label_file_text = "Run \Directory\n to be cleaned"
        self.label_file_dir = tkinter.Label(self.root, text=label_file_text)
        # grid() assigns locations. sticky moves it NSE or W within cell
        self.label_file_dir.grid(row=fd_row, column=fd_column,
                                 sticky=tkinter.E)
        self.entrytext_file = tkinter.StringVar()
        # Input for data
        self.e_file_dir = tkinter.Entry(
                        self.root,
                        width=95,
                        textvariable=self.entrytext_file)
        self.e_file_dir.grid(row=fd_row, column=fd_column+1,
                             columnspan=10, sticky=tkinter.W)

        expr_row = 3
        expr_column = 0

        self.label_stressor = tkinter.Label(self.root, text="Stressors")
        self.label_stressor.grid(row=expr_row, column=expr_column,
                                 sticky=tkinter.E)
        self.entrytext_stressor = tkinter.StringVar()
        self.e_stressor = tkinter.Entry(
                        self.root,
                        width=15,
                        textvariable=self.entrytext_stressor)
        self.e_stressor.grid(row=expr_row, column=expr_column+1,
                             sticky=tkinter.W)
        self.e_stressor.insert(0, "W,D")

        self.label_species = tkinter.Label(self.root, text="Species")
        self.label_species.grid(row=expr_row, column=expr_column+1,
                                sticky=tkinter.E)
        self.entrytext_species = tkinter.StringVar()
        self.e_species = tkinter.Entry(
                       self.root,
                       width=15,
                       textvariable=self.entrytext_species)
        self.e_species.grid(row=expr_row, column=expr_column+2,
                            sticky=tkinter.W)
        self.e_species.insert(0, "J,N,C")

        filler_1_text = "                "
        self.label_run_1 = tkinter.Label(self.root, text=filler_1_text)
        self.label_run_1.grid(row=5, column=0)

        run_row = 6
        run_column = 0

        cc_toggle_text = u"Toggle CC_plot"
        self.var_cc_toggle = tkinter.IntVar()
        self.checkbutton_cc_toggle = tkinter.Checkbutton(
                                   self.root,
                                   text=cc_toggle_text,
                                   variable=self.var_cc_toggle,
                                   selectcolor="gray55",
                                   command=self.toggle_cc,
                                   bg='gray85')
        self.checkbutton_cc_toggle.grid(row=run_row, column=run_column,
                                        ipadx=250, columnspan=3)

        irt_toggle_text = "Toggle IRT_plot"
        self.var_irt_toggle = tkinter.IntVar()
        self.checkbutton_irt_toggle = tkinter.Checkbutton(
                                    self.root,
                                    text=irt_toggle_text,
                                    variable=self.var_irt_toggle,
                                    selectcolor="gray55",
                                    command=self.toggle_irt,
                                    bg='gray85')
        self.checkbutton_irt_toggle.grid(row=run_row, column=run_column+3,
                                         ipadx=250, columnspan=3)

        range_row = 7
        range_column = 0
        self.label_range_c = tkinter.Label(self.root, text="Range")
        self.label_range_c.grid(row=range_row, column=range_column,
                                sticky=tkinter.E)
        self.entrytext_range_c = tkinter.StringVar()
        self.e_range_c = tkinter.Entry(
                       self.root,
                       width=5,
                       textvariable=self.entrytext_range_c)
        self.e_range_c.grid(row=range_row, column=range_column+1,
                            sticky=tkinter.W)
        self.e_range_c.insert(0, "0,1")

        self.var_r_toggle_c = tkinter.IntVar()
        self.checkbutton_r_toggle_c = tkinter.Checkbutton(
                                    self.root,
                                    text="Don't Define Range",
                                    variable=self.var_r_toggle_c,
                                    selectcolor="gray",
                                    command=self.toggle_range_c)
        self.checkbutton_r_toggle_c.grid(row=range_row, column=range_column+2,
                                         sticky=tkinter.W)

        self.label_range_i = tkinter.Label(self.root, text="Range")
        self.label_range_i.grid(row=range_row, column=range_column+3,
                                sticky=tkinter.E)
        self.entrytext_range_i = tkinter.StringVar()
        self.e_range_i = tkinter.Entry(
                       self.root,
                       width=5,
                       textvariable=self.entrytext_range_i)
        self.e_range_i.grid(row=range_row, column=range_column+4,
                            sticky=tkinter.W)
        self.e_range_i.insert(0, "0,1")

        self.var_r_toggle_i = tkinter.IntVar()
        self.checkbutton_r_toggle_i = tkinter.Checkbutton(
                                    self.root,
                                    text="Don't Define Range",
                                    variable=self.var_r_toggle_i,
                                    selectcolor="gray",
                                    command=self.toggle_range_i)
        self.checkbutton_r_toggle_i.grid(row=range_row, column=range_column+5,
                                         sticky=tkinter.W)

        plot_std_row = 8
        plot_std_column = 0
        plot_std_text_c = "Set SD Plot Range Scalar"
        self.label_plot_std_c = tkinter.Label(self.root, text=plot_std_text_c)
        self.label_plot_std_c.grid(row=plot_std_row, column=plot_std_column,
                                   sticky=tkinter.E)
        self.entrytext_plot_std_c = tkinter.StringVar()
        self.e_plot_std_c = tkinter.Entry(
                          self.root,
                          width=15,
                          textvariable=self.entrytext_plot_std_c)
        self.e_plot_std_c.grid(row=plot_std_row, column=plot_std_column+1,
                               sticky=tkinter.SW)
        self.e_plot_std_c.insert(0, " ")

        self.var_pstd_toggle_c = tkinter.IntVar()

        pstd_toggle_text_c = u"Toggle Plot Standard Deviation"
        self.checkbutton_pstd_toggle_c = tkinter.Checkbutton(
                                       self.root,
                                       text=pstd_toggle_text_c,
                                       variable=self.var_pstd_toggle_c,
                                       selectcolor="gray",
                                       command=self.toggle_plot_std_c)
        self.checkbutton_pstd_toggle_c.grid(row=plot_std_row,
                                            column=plot_std_column+2,
                                            sticky=tkinter.W)

        plot_std_text_i = "Set SD Plot Range Scalar"
        self.label_plot_std_i = tkinter.Label(self.root, text=plot_std_text_i)
        self.label_plot_std_i.grid(row=plot_std_row, column=plot_std_column+3,
                                   sticky=tkinter.E)
        self.entrytext_plot_std_i = tkinter.StringVar()
        self.e_plot_std_i = tkinter.Entry(
                          self.root,
                          width=15,
                          textvariable=self.entrytext_plot_std_i)
        self.e_plot_std_i.grid(row=plot_std_row, column=plot_std_column+4,
                               sticky=tkinter.SW)
        self.e_plot_std_i.insert(0, " ")

        self.var_pstd_toggle_i = tkinter.IntVar()

        pstd_toggle_text_i = u"Toggle Plot Standard Deviation"
        self.checkbutton_pstd_toggle_i = tkinter.Checkbutton(
                                       self.root,
                                       text=pstd_toggle_text_i,
                                       variable=self.var_pstd_toggle_i,
                                       selectcolor="gray",
                                       command=self.toggle_plot_std_i)
        self.checkbutton_pstd_toggle_i.grid(row=plot_std_row,
                                            column=plot_std_column+5,
                                            sticky=tkinter.W)

        add_on_row = 9
        add_on_column = 0
        label_add_on_text = "Stat Add-ons"
        self.label_add_on = tkinter.Label(self.root, text=label_add_on_text)
        # grid() assigns locations. sticky moves it NSE or W within cell
        self.label_add_on.grid(row=add_on_row, column=add_on_column,
                               sticky=tkinter.E, rowspan=2)

        self.var_zeros_c = tkinter.IntVar()
        self.checkbutton_zeros_c = tkinter.Checkbutton(
                                 self.root,
                                 text="Remove Zeros",
                                 variable=self.var_zeros_c,
                                 selectcolor="gray")
        self.checkbutton_zeros_c.grid(row=add_on_row, column=add_on_column+1,
                                      sticky=tkinter.W)

        self.var_raw_c = tkinter.IntVar()
        raw_text_c = "Create Plots"
        self.checkbutton_raw_c = tkinter.Checkbutton(
                               self.root,
                               text=raw_text_c,
                               variable=self.var_raw_c,
                               selectcolor="gray")
        self.checkbutton_raw_c.grid(row=add_on_row, column=add_on_column+2,
                                    sticky=tkinter.W)

        self.var_reanalyze_c = tkinter.IntVar()
        self.checkbutton_reanalyze_c = tkinter.Checkbutton(
                                     self.root,
                                     text="Reanalyze Stat",
                                     variable=self.var_reanalyze_c,
                                     selectcolor="gray")
        self.checkbutton_reanalyze_c.grid(row=add_on_row+1,
                                          column=add_on_column+1,
                                          sticky=tkinter.W)

        self.var_zeros_i = tkinter.IntVar()
        self.checkbutton_zeros_i = tkinter.Checkbutton(
                                 self.root,
                                 text="Remove Zeros",
                                 variable=self.var_zeros_i,
                                 selectcolor="gray")
        self.checkbutton_zeros_i.grid(row=add_on_row, column=add_on_column+4,
                                      sticky=tkinter.W)

        self.var_raw_i = tkinter.IntVar()
        raw_text_i = u"Create Plots"
        self.checkbutton_raw_i = tkinter.Checkbutton(
                               self.root,
                               text=raw_text_i,
                               variable=self.var_raw_i,
                               selectcolor="gray")
        self.checkbutton_raw_i.grid(row=add_on_row, column=add_on_column+5,
                                    sticky=tkinter.W)

        self.var_reanalyze_i = tkinter.IntVar()
        self.checkbutton_reanalyze_i = tkinter.Checkbutton(
                                     self.root,
                                     text="Reanalyze Stat",
                                     variable=self.var_reanalyze_i,
                                     selectcolor="gray")
        self.checkbutton_reanalyze_i.grid(row=add_on_row+1,
                                          column=add_on_column+4,
                                          sticky=tkinter.W)

        filler_2_text = "      "
        self.label_run_2 = tkinter.Label(self.root, text=filler_2_text)
        self.label_run_2.grid(row=11, column=0, sticky=tkinter.E)

        run_1_row = 12
        run_1_column = 0
        honey_toggle_text = "Toggle Honey_plot"
        self.var_honey_toggle = tkinter.IntVar()
        self.checkbutton_honey_toggle = tkinter.Checkbutton(
                                      self.root,
                                      text=honey_toggle_text,
                                      variable=self.var_honey_toggle,
                                      selectcolor="gray55",
                                      command=self.toggle_honey,
                                      bg='gray85')
        self.checkbutton_honey_toggle.grid(row=run_1_row, column=run_1_column,
                                           ipadx=250, columnspan=3)

        pep_toggle_text = "Toggle Pep_plot"
        self.var_pep_toggle = tkinter.IntVar()
        self.checkbutton_pep_toggle = tkinter.Checkbutton(
                                    self.root,
                                    text=pep_toggle_text,
                                    variable=self.var_pep_toggle,
                                    selectcolor="gray55",
                                    command=self.toggle_pep,
                                    bg='gray85')
        self.checkbutton_pep_toggle.grid(row=run_1_row, column=run_1_column+3,
                                         ipadx=250, columnspan=3)

        range_1_row = 13
        range_1_column = 0
        self.label_range_h = tkinter.Label(self.root, text="Range")
        self.label_range_h.grid(row=range_1_row, column=range_1_column,
                                sticky=tkinter.E)
        self.entrytext_range_h = tkinter.StringVar()
        self.e_range_h = tkinter.Entry(
                       self.root,
                       width=5,
                       textvariable=self.entrytext_range_h)
        self.e_range_h.grid(row=range_1_row, column=range_1_column+1,
                            sticky=tkinter.W)
        self.e_range_h.insert(0, "0,1")

        self.var_r_toggle_h = tkinter.IntVar()
        self.checkbutton_r_toggle_h = tkinter.Checkbutton(
                                    self.root,
                                    text="Don't Define Range",
                                    variable=self.var_r_toggle_h,
                                    selectcolor="gray",
                                    command=self.toggle_range_h)
        self.checkbutton_r_toggle_h.grid(row=range_1_row,
                                         column=range_1_column+2,
                                         sticky=tkinter.W)

        self.label_range_p = tkinter.Label(self.root, text="Range")
        self.label_range_p.grid(row=range_1_row, column=range_1_column+3,
                                sticky=tkinter.E)
        self.entrytext_range_p = tkinter.StringVar()
        self.e_range_p = tkinter.Entry(
                       self.root,
                       width=5,
                       textvariable=self.entrytext_range_p)
        self.e_range_p.grid(row=range_1_row, column=range_1_column+4,
                            sticky=tkinter.W)
        self.e_range_p.insert(0, "0,1")

        self.var_r_toggle_p = tkinter.IntVar()
        self.checkbutton_r_toggle_p = tkinter.Checkbutton(
                                    self.root,
                                    text="Don't Define Range",
                                    variable=self.var_r_toggle_p,
                                    selectcolor="gray",
                                    command=self.toggle_range_p)
        self.checkbutton_r_toggle_p.grid(row=range_1_row,
                                         column=range_1_column+5,
                                         sticky=tkinter.W)

        plot_std_1_row = 14
        plot_std_1_column = 0
        plot_std_text_h = "Set SD Plot Range Scalar"
        self.label_plot_std_h = tkinter.Label(self.root, text=plot_std_text_h)
        self.label_plot_std_h.grid(row=plot_std_1_row,
                                   column=plot_std_1_column,
                                   sticky=tkinter.E)
        self.entrytext_plot_std_h = tkinter.StringVar()
        self.e_plot_std_h = tkinter.Entry(
                          self.root,
                          width=15,
                          textvariable=self.entrytext_plot_std_h)
        self.e_plot_std_h.grid(row=plot_std_1_row, column=plot_std_1_column+1,
                               sticky=tkinter.SW)
        self.e_plot_std_h.insert(0, " ")

        self.var_pstd_toggle_h = tkinter.IntVar()

        pstd_toggle_text_h = u"Toggle Plot Standard Deviation"
        self.checkbutton_pstd_toggle_h = tkinter.Checkbutton(
                                       self.root,
                                       text=pstd_toggle_text_h,
                                       variable=self.var_pstd_toggle_h,
                                       selectcolor="gray",
                                       command=self.toggle_plot_std_h)
        self.checkbutton_pstd_toggle_h.grid(row=plot_std_1_row,
                                            column=plot_std_1_column+2,
                                            sticky=tkinter.W)

        plot_std_text_p = "Set SD Plot Range Scalar"
        self.label_plot_std_p = tkinter.Label(self.root, text=plot_std_text_p)
        self.label_plot_std_p.grid(row=plot_std_1_row,
                                   column=plot_std_1_column+3,
                                   sticky=tkinter.E)
        self.entrytext_plot_std_p = tkinter.StringVar()
        self.e_plot_std_p = tkinter.Entry(
                          self.root,
                          width=15,
                          textvariable=self.entrytext_plot_std_p)
        self.e_plot_std_p.grid(row=plot_std_1_row, column=plot_std_1_column+4,
                               sticky=tkinter.SW)
        self.e_plot_std_p.insert(0, " ")

        self.var_pstd_toggle_p = tkinter.IntVar()

        pstd_toggle_text_p = u"Toggle Plot Standard Deviation"
        self.checkbutton_pstd_toggle_p = tkinter.Checkbutton(
                                       self.root,
                                       text=pstd_toggle_text_p,
                                       variable=self.var_pstd_toggle_p,
                                       selectcolor="gray",
                                       command=self.toggle_plot_std_p)
        self.checkbutton_pstd_toggle_p.grid(row=plot_std_1_row,
                                            column=plot_std_1_column+5,
                                            sticky=tkinter.W)

        add_on_1_row = 15
        add_on_1_column = 0
        label_add_on_1_text = "Stat Add-ons"
        self.label_add_on_1 = tkinter.Label(
                            self.root,
                            text=label_add_on_1_text)
        self.label_add_on_1.grid(row=add_on_1_row, column=add_on_1_column,
                                 sticky=tkinter.E, rowspan=2)

        self.var_zeros_h = tkinter.IntVar()
        self.checkbutton_zeros_h = tkinter.Checkbutton(
                                 self.root,
                                 text="Remove Zeros",
                                 variable=self.var_zeros_h,
                                 selectcolor="gray")
        self.checkbutton_zeros_h.grid(row=add_on_1_row,
                                      column=add_on_1_column+1,
                                      sticky=tkinter.W)

        self.var_raw_h = tkinter.IntVar()
        raw_text_h = "Create Plots"
        self.checkbutton_raw_h = tkinter.Checkbutton(
                               self.root,
                               text=raw_text_h,
                               variable=self.var_raw_h,
                               selectcolor="gray")
        self.checkbutton_raw_h.grid(row=add_on_1_row, column=add_on_1_column+2,
                                    sticky=tkinter.W)

        self.var_reanalyze_h = tkinter.IntVar()
        self.checkbutton_reanalyze_h = tkinter.Checkbutton(
                                     self.root,
                                     text="Reanalyze Stat",
                                     variable=self.var_reanalyze_h,
                                     selectcolor="gray")
        self.checkbutton_reanalyze_h.grid(row=add_on_1_row+1,
                                          column=add_on_1_column+1,
                                          sticky=tkinter.W)

        self.var_zeros_p = tkinter.IntVar()
        self.checkbutton_zeros_p = tkinter.Checkbutton(
                                 self.root,
                                 text="Remove Zeros",
                                 variable=self.var_zeros_p,
                                 selectcolor="gray")
        self.checkbutton_zeros_p.grid(row=add_on_1_row,
                                      column=add_on_1_column+4,
                                      sticky=tkinter.W)

        self.var_raw_p = tkinter.IntVar()
        raw_text_p = u"Create Plots"
        self.checkbutton_raw_p = tkinter.Checkbutton(
                               self.root,
                               text=raw_text_p,
                               variable=self.var_raw_p,
                               selectcolor="gray")
        self.checkbutton_raw_p.grid(row=add_on_1_row, column=add_on_1_column+5,
                                    sticky=tkinter.W)

        self.var_reanalyze_p = tkinter.IntVar()
        self.checkbutton_reanalyze_p = tkinter.Checkbutton(
                                     self.root,
                                     text="Reanalyze Stat",
                                     variable=self.var_reanalyze_p,
                                     selectcolor="gray")
        self.checkbutton_reanalyze_p.grid(row=add_on_1_row+1,
                                          column=add_on_1_column+4,
                                          sticky=tkinter.W)

        filler_3_text = "                "
        self.label_run_3 = tkinter.Label(self.root, text=filler_3_text)
        self.label_run_3.grid(row=add_on_1_row+3, column=0)

        button_row = 19
        button_column = 2
        # Creates a button widget for quitting the window
        self.buttontext_kill = tkinter.StringVar()
        self.buttontext_kill.set("Kill Program")
        self.button_kill = tkinter.Button(
                         self.root,
                         textvariable=self.buttontext_kill,
                         command=self.quit_click)
        self.button_kill.configure(bg='IndianRed')  # or firebrick2
        self.button_kill.grid(row=button_row, column=button_column,
                              ipadx=21, ipady=5)

        # Creates a button widget for inserting inputs to program
        self.buttontext_previous = tkinter.StringVar()
        self.buttontext_previous.set("Previous Entries")
        self.button_previous = tkinter.Button(
                             self.root,
                             textvariable=self.buttontext_previous,
                             command=self.fill_recent_entries)
        self.button_previous.configure(bg='deep sky blue2')
        self.button_previous.grid(row=button_row, column=button_column+1,
                                  ipadx=35, ipady=5, sticky=tkinter.E)

        # Creates a button widget for submitting inputs to program
        self.buttontext_submit = tkinter.StringVar()
        self.buttontext_submit.set("Submit")
        self.button_submit = tkinter.Button(
                           self.root,
                           textvariable=self.buttontext_submit,
                           command=self.write_string_to_file)
        # gives button a background color
        self.button_submit.configure(bg='Spring Green2')  # or chartruese2
        self.button_submit.grid(row=button_row, column=button_column+3,
                                ipadx=35, ipady=5, sticky=tkinter.W)

        # These are for keyboard execution of the GUI
        # if these are to be deleted, be sure to delete
        # 'event=None' from defs
        # (keystroke, command)
        self.root.bind("<Down>", self.fill_recent_entries)
        self.root.bind("<Return>", self.write_string_to_file)
        self.root.bind("<Escape>", self.quit_click)

        # enables window to appear
        self.root.mainloop()

        return

    def toggle_cc(self):
        """ When called, it inserts the string into the entry box"""

        if int(self.var_cc_toggle.get()) == 0:
            self.checkbutton_cc_toggle.configure(bg='gray85')

        if int(self.var_cc_toggle.get()) == 1:
            self.checkbutton_cc_toggle.configure(bg='spring green2')

        return

    def toggle_irt(self):
        """ When called, it inserts the string into the entry box"""

        if int(self.var_irt_toggle.get()) == 0:
            self.checkbutton_irt_toggle.configure(bg='gray85')

        if int(self.var_irt_toggle.get()) == 1:
            self.checkbutton_irt_toggle.configure(bg='spring green2')

        return

    def toggle_honey(self):
        """ When called, it inserts the string into the entry box"""

        if int(self.var_honey_toggle.get()) == 0:
            self.checkbutton_honey_toggle.configure(bg='gray85')

        if int(self.var_honey_toggle.get()) == 1:
            self.checkbutton_honey_toggle.configure(bg='spring green2')

        return

    def toggle_pep(self):
        """ When called, it inserts the string into the entry box"""

        if int(self.var_pep_toggle.get()) == 0:
            self.checkbutton_pep_toggle.configure(bg='gray85')

        if int(self.var_pep_toggle.get()) == 1:
            self.checkbutton_pep_toggle.configure(bg='spring green2')

        return

    def toggle_range_c(self):
        """ When called, it inserts the string into the entry box"""

        if int(self.var_r_toggle_c.get()) == 0:
            self.e_range_c.delete(0, tkinter.END)
            self.e_range_c.insert(0, "0,1")

        if int(self.var_r_toggle_c.get()) == 1:
            self.e_range_c.delete(0, tkinter.END)
            self.e_range_c.insert(0, "")

        return

    def toggle_range_i(self):
        """ When called, it inserts the string into the entry box"""

        if int(self.var_r_toggle_i.get()) == 0:
            self.e_range_i.delete(0, tkinter.END)
            self.e_range_i.insert(0, "0,1")

        if int(self.var_r_toggle_i.get()) == 1:
            self.e_range_i.delete(0, tkinter.END)
            self.e_range_i.insert(0, "")

        return

    def toggle_range_h(self):
        """ When called, it inserts the string into the entry box"""

        if int(self.var_r_toggle_h.get()) == 0:
            self.e_range_h.delete(0, tkinter.END)
            self.e_range_h.insert(0, "0,10")

        if int(self.var_r_toggle_h.get()) == 1:
            self.e_range_h.delete(0, tkinter.END)
            self.e_range_h.insert(0, "")

        return

    def toggle_range_p(self):
        """ When called, it inserts the string into the entry box"""

        if int(self.var_r_toggle_p.get()) == 0:
            self.e_range_p.delete(0, tkinter.END)
            self.e_range_p.insert(0, "0,10")

        if int(self.var_r_toggle_p.get()) == 1:
            self.e_range_p.delete(0, tkinter.END)
            self.e_range_p.insert(0, "")

        return

    def toggle_plot_std_c(self):
        """ When called, it inserts the string into the entry box"""

        if int(self.var_pstd_toggle_c.get()) == 0:
            self.e_plot_std_c.delete(0, tkinter.END)
            self.e_plot_std_c.insert(0, " ")

        if int(self.var_pstd_toggle_c.get()) == 1:
            self.e_plot_std_c.delete(0, tkinter.END)
            self.e_plot_std_c.insert(0, "1.5")

        return

    def toggle_plot_std_i(self):
        """ When called, it inserts the string into the entry box"""

        if int(self.var_pstd_toggle_i.get()) == 0:
            self.e_plot_std_i.delete(0, tkinter.END)
            self.e_plot_std_i.insert(0, " ")

        if int(self.var_pstd_toggle_i.get()) == 1:
            self.e_plot_std_i.delete(0, tkinter.END)
            self.e_plot_std_i.insert(0, "1.5")

        return

    def toggle_plot_std_h(self):
        """ When called, it inserts the string into the entry box"""

        if int(self.var_pstd_toggle_h.get()) == 0:
            self.e_plot_std_h.delete(0, tkinter.END)
            self.e_plot_std_h.insert(0, " ")

        if int(self.var_pstd_toggle_h.get()) == 1:
            self.e_plot_std_h.delete(0, tkinter.END)
            self.e_plot_std_h.insert(0, "1.5")

        return

    def toggle_plot_std_p(self):
        """ When called, it inserts the string into the entry box"""

        if int(self.var_pstd_toggle_p.get()) == 0:
            self.e_plot_std_p.delete(0, tkinter.END)
            self.e_plot_std_p.insert(0, " ")

        if int(self.var_pstd_toggle_p.get()) == 1:
            self.e_plot_std_p.delete(0, tkinter.END)
            self.e_plot_std_p.insert(0, "1.5")

        return

    def write_string_to_file(self, event=None):
        """Command for button widget

        for given inputs, write them into a new text file and then close it
        """

        # make a list of each entry that was in the text box GUI
        entered_parameters = [self.entrytext_file.get(),
                              self.entrytext_stressor.get(),
                              self.entrytext_species.get(),

                              self.var_cc_toggle.get(),
                              self.entrytext_range_c.get(),
                              self.var_r_toggle_c.get(),
                              self.entrytext_plot_std_c.get(),
                              self.var_pstd_toggle_c.get(),
                              self.var_zeros_c.get(),
                              self.var_raw_c.get(),
                              self.var_reanalyze_c.get(),

                              self.var_irt_toggle.get(),
                              self.entrytext_range_i.get(),
                              self.var_r_toggle_i.get(),
                              self.entrytext_plot_std_i.get(),
                              self.var_pstd_toggle_i.get(),
                              self.var_zeros_i.get(),
                              self.var_raw_i.get(),
                              self.var_reanalyze_i.get(),

                              self.var_honey_toggle.get(),
                              self.entrytext_range_h.get(),
                              self.var_r_toggle_h.get(),
                              self.entrytext_plot_std_h.get(),
                              self.var_pstd_toggle_h.get(),
                              self.var_zeros_h.get(),
                              self.var_raw_h.get(),
                              self.var_reanalyze_h.get(),

                              self.var_pep_toggle.get(),
                              self.entrytext_range_p.get(),
                              self.var_r_toggle_p.get(),
                              self.entrytext_plot_std_p.get(),
                              self.var_pstd_toggle_p.get(),
                              self.var_zeros_p.get(),
                              self.var_raw_p.get(),
                              self.var_reanalyze_p.get()]

        # define a file name, without .txt
        file_name = "USE_Cleaner"

        # launch function to create and write to the file with the inputs
        Create_And_Write_File(file_name, entered_parameters)

        # close GUI window
        self.root.destroy()

        return

    def fill_recent_entries(self, event=None):
        """Command (to autofill) for a button widget

        this is to make users life easier with an autofill option
        by entering the last parameters the user gave into Entry widgets
        -only works if this program has been run before
        """

        # previously made file
        entry_file = "USE_Cleaner.txt"

        # if file exists, complete the autofill
        if os.path.exists(entry_file):

            # obtain parameters from file
            user_inputs = Read_Text_File(entry_file)

            # redefine inputs
            file_name = user_inputs[0]
            stressors = user_inputs[1]
            species = user_inputs[2]

            var_cc_toggle = user_inputs[3]
            text_range_c = user_inputs[4]
            var_r_toggle_c = user_inputs[5]
            text_plot_std_c = user_inputs[6]
            var_pstd_toggle_c = user_inputs[7]
            var_zeros_c = user_inputs[8]
            var_raw_c = user_inputs[9]
            var_reanalyze_c = user_inputs[10]

            var_irt_toggle = user_inputs[11]
            text_range_i = user_inputs[12]
            var_r_toggle_i = user_inputs[13]
            text_plot_std_i = user_inputs[14]
            var_pstd_toggle_i = user_inputs[15]
            var_zeros_i = user_inputs[16]
            var_raw_i = user_inputs[17]
            var_reanalyze_i = user_inputs[18]

            var_honey_toggle = user_inputs[19]
            text_range_h = user_inputs[20]
            var_r_toggle_h = user_inputs[21]
            text_plot_std_h = user_inputs[22]
            var_pstd_toggle_h = user_inputs[23]
            var_zeros_h = user_inputs[24]
            var_raw_h = user_inputs[25]
            var_reanalyze_h = user_inputs[26]

            var_pep_toggle = user_inputs[27]
            text_range_p = user_inputs[28]
            var_r_toggle_p = user_inputs[29]
            text_plot_std_p = user_inputs[30]
            var_pstd_toggle_p = user_inputs[31]
            var_zeros_p = user_inputs[32]
            var_raw_p = user_inputs[33]
            var_reanalyze_p = user_inputs[34]

            # insert inputs into Entry boxes in the GUI
            # (skipped characters, variable)
            self.e_file_dir.delete(0, tkinter.END)
            self.e_file_dir.insert(0, file_name)

            self.e_stressor.delete(0, tkinter.END)
            self.e_stressor.insert(0, stressors)

            self.e_species.delete(0, tkinter.END)
            self.e_species.insert(0, species)

            # select() chooses the box to be switched on.
            if int(var_cc_toggle) == 1:
                self.checkbutton_cc_toggle.select()
                self.toggle_cc()
            self.e_range_c.delete(0, tkinter.END)
            self.e_range_c.insert(0, text_range_c)
            if int(var_r_toggle_c) == 1:
                self.checkbutton_r_toggle_c.select()
            self.e_plot_std_c.delete(0, tkinter.END)
            self.e_plot_std_c.insert(0, text_plot_std_c)
            if int(var_pstd_toggle_c) == 1:
                self.checkbutton_pstd_toggle_c.select()
            if int(var_zeros_c) == 1:
                self.checkbutton_zeros_c.select()
            if int(var_raw_c) == 1:
                self.checkbutton_raw_c.select()
            if int(var_reanalyze_c) == 1:
                self.checkbutton_reanalyze_c.select()

            if int(var_irt_toggle) == 1:
                self.checkbutton_irt_toggle.select()
                self.toggle_irt()
            self.e_range_i.delete(0, tkinter.END)
            self.e_range_i.insert(0, text_range_i)
            if int(var_r_toggle_i) == 1:
                self.checkbutton_r_toggle_i.select()
            self.e_plot_std_i.delete(0, tkinter.END)
            self.e_plot_std_i.insert(0, text_plot_std_i)
            if int(var_pstd_toggle_i) == 1:
                self.checkbutton_pstd_toggle_i.select()
            if int(var_zeros_i) == 1:
                self.checkbutton_zeros_i.select()
            if int(var_raw_i) == 1:
                self.checkbutton_raw_i.select()
            if int(var_reanalyze_i) == 1:
                self.checkbutton_reanalyze_i.select()

            if int(var_honey_toggle) == 1:
                self.checkbutton_honey_toggle.select()
                self.toggle_honey()
            self.e_range_h.delete(0, tkinter.END)
            self.e_range_h.insert(0, text_range_h)
            if int(var_r_toggle_h) == 1:
                self.checkbutton_r_toggle_h.select()
            self.e_plot_std_h.delete(0, tkinter.END)
            self.e_plot_std_h.insert(0, text_plot_std_h)
            if int(var_pstd_toggle_h) == 1:
                self.checkbutton_pstd_toggle_h.select()
            if int(var_zeros_h) == 1:
                self.checkbutton_zeros_h.select()
            if int(var_raw_h) == 1:
                self.checkbutton_raw_h.select()
            if int(var_reanalyze_h) == 1:
                self.checkbutton_reanalyze_h.select()

            if int(var_pep_toggle) == 1:
                self.checkbutton_pep_toggle.select()
                self.toggle_pep()
            self.e_range_p.delete(0, tkinter.END)
            self.e_range_p.insert(0, text_range_p)
            if int(var_r_toggle_p) == 1:
                self.checkbutton_r_toggle_p.select()
            self.e_plot_std_p.delete(0, tkinter.END)
            self.e_plot_std_p.insert(0, text_plot_std_p)
            if int(var_pstd_toggle_p) == 1:
                self.checkbutton_pstd_toggle_p.select()
            if int(var_zeros_p) == 1:
                self.checkbutton_zeros_p.select()
            if int(var_raw_p) == 1:
                self.checkbutton_raw_p.select()
            if int(var_reanalyze_p) == 1:
                self.checkbutton_reanalyze_p.select()

        else:
            print(entry_file + " does not exist! Enter parameters by hand")

        return

    def quit_click(self, event=None):
        """To handle quiting the window we created"""

        self.root.destroy()

        # User decides to kill program and then the scripts exits safely
        print("\nXXXXXXXXXXXX "
              "Program successfully killed by User "
              "XXXXXXXXXXXX\n")
        sys.exit(1)

        return


def Create_And_Write_File(file_name, data):
    """Creates and writes to a new file

    file_name -> str
    data -> list
    """

    # file_name is usually given without a type so we make it .txt
    full_file_name = file_name + ".txt"

    # create new file with full file name
    new_file = open(full_file_name, 'w+')

    # write data to the file
    for i in range(len(data)):
        new_file.write(str(data[i]) + "\n")

    # close file
    new_file.close()

    return


def Read_Text_File(file_name):
    """Read in a text file

    file_name -> str
    """

    # open the file
    x = open(file_name, 'rU')

    # read the contents and assign into list
    data_in_file = x.readlines()

    # initialize list for the forloop
    contents = []

    # strip data then append it into initialized list
    for i in range(len(data_in_file)):
        stripped_data = data_in_file[i].strip("\n")
        contents.append(stripped_data)

    # close file
    x.close()

    return contents


def Clean_CC_Plot(clean_csv, file_name, zeros_toggle,
                  custom_range, range_onoff):
    """This function  reads in the csv from user and returns cleaned data

    We assign corresponding data to columns. Then adjust the data to fit within
    a certain range. if out of bounds, reassign a nan value. Then write clean
    data to new csv.
    clean_csv -> str
    file_name -> str
    zeros_toggle -> int
    custon_range -> list of floats
    range_onoff -> int
    """

    print("***\n**\n*\nLoading in the raw data from " + file_name + "\n.\n.")

    data = np.genfromtxt(file_name, dtype=None, delimiter=",",
                         names=True, skip_header=0)

    cire = data['CIRE'].astype(np.float)
    mtci = data['MTCI'].astype(np.float)
    ccci = data['CCCI'].astype(np.float)
    cccia = data['CCCIA'].astype(np.float)
    ci800 = data['CI800'].astype(np.float)
    datt = data['DATT'].astype(np.float)
    datta = data['DATTA'].astype(np.float)
    ndare = data['NDARE'].astype(np.float)
    ndre = data['NDRE'].astype(np.float)
    ndrre = data['NDRRE'].astype(np.float)
    ndvia = data['NDVIA'].astype(np.float)
    ndvir = data['NDVIR'].astype(np.float)
    nvg2 = data['NVG2'].astype(np.float)
    nvg800 = data['NVG800'].astype(np.float)
    pri = data['PRI'].astype(np.float)

    # some lines have all nan values and it
    # throws a warning about Runtime.
    # Nothing to worry about, just numpy at the edges
    # of its formal structure
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", category=RuntimeWarning)

        # this means the range_toggle is off and we should include the range.
        if range_onoff == 0:

            c_min = custom_range[0]
            c_max = custom_range[1]
            print("Removing values outside " + str(c_min) +
                  " and " + str(c_max))

            # data outside of range gets nan value
            cire[(cire < c_min) | (cire > c_max)] = np.nan
            mtci[(mtci < c_min) | (mtci > c_max)] = np.nan
            ccci[(ccci < c_min) | (ccci > c_max)] = np.nan
            cccia[(cccia < c_min) | (cccia > c_max)] = np.nan
            ci800[(ci800 < c_min) | (ci800 > c_max)] = np.nan
            datt[(datt < c_min) | (datt > c_max)] = np.nan
            datta[(datta < c_min) | (datta > c_max)] = np.nan
            ndare[(ndare < c_min) | (ndare > c_max)] = np.nan
            ndre[(ndre < c_min) | (ndre > c_max)] = np.nan
            ndrre[(ndrre < c_min) | (ndrre > c_max)] = np.nan
            ndvia[(ndvia < c_min) | (ndvia > c_max)] = np.nan
            ndvir[(ndvir < c_min) | (ndvir > c_max)] = np.nan
            nvg2[(nvg2 < c_min) | (nvg2 > c_max)] = np.nan
            nvg800[(nvg800 < c_min) | (nvg800 > c_max)] = np.nan
            pri[(pri < c_min) | (pri > c_max)] = np.nan

        # If zero toggle is on then replace zeros with nan
        if zeros_toggle == 1:
            print("Zeros being eliminated")
            # for elements in array,
            # replace values outside of the range with nan
            cire[(cire == 0.)] = np.nan
            mtci[(mtci == 0.)] = np.nan
            ccci[(ccci == 0.)] = np.nan
            cccia[(cccia == 0.)] = np.nan
            ci800[(ci800 == 0.)] = np.nan
            datt[(datt == 0.)] = np.nan
            datta[(datta == 0.)] = np.nan
            ndare[(ndare == 0.)] = np.nan
            ndre[(ndre == 0.)] = np.nan
            ndrre[(ndrre == 0.)] = np.nan
            ndvia[(ndvia == 0.)] = np.nan
            ndvir[(ndvir == 0.)] = np.nan
            nvg2[(nvg2 == 0.)] = np.nan
            nvg800[(nvg800 == 0.)] = np.nan
            pri[(pri == 0.)] = np.nan

    # write untouched val and modified array to new file
    with open(clean_csv, 'wb') as cc:
        cleaned = csv.writer(cc, delimiter=",")

        # i marks the current row in spreadsheet
        for i in range(len(ccci)):

            if i == 0:
                cleaned.writerow(data.dtype.names)
                print("Replacing numbers out of range")

            cleaned_data = data[i]

            # j marks element in each row
            for j in range(len(data.dtype.names)):

                # replace out of bounds data with nan
                if str(data.dtype.names[j]) == "CIRE":
                    cleaned_data[j] = cire[i]
                elif str(data.dtype.names[j]) == "MTCI":
                    cleaned_data[j] = mtci[i]
                elif str(data.dtype.names[j]) == "CCCI":
                    cleaned_data[j] = ccci[i]
                elif str(data.dtype.names[j]) == "CCCIA":
                    cleaned_data[j] = cccia[i]
                elif str(data.dtype.names[j]) == "CI800":
                    cleaned_data[j] = ci800[i]
                elif str(data.dtype.names[j]) == "DATT":
                    cleaned_data[j] = datt[i]
                elif str(data.dtype.names[j]) == "DATTA":
                    cleaned_data[j] = datta[i]
                elif str(data.dtype.names[j]) == "NDARE":
                    cleaned_data[j] = ndare[i]
                elif str(data.dtype.names[j]) == "NDRE":
                    cleaned_data[j] = ndre[i]
                elif str(data.dtype.names[j]) == "NDRRE":
                    cleaned_data[j] = ndrre[i]
                elif str(data.dtype.names[j]) == "NDVIA":
                    cleaned_data[j] = ndvia[i]
                elif str(data.dtype.names[j]) == "NDVIR":
                    cleaned_data[j] = ndvir[i]
                elif str(data.dtype.names[j]) == "NVG2":
                    cleaned_data[j] = nvg2[i]
                elif str(data.dtype.names[j]) == "NVG800":
                    cleaned_data[j] = nvg800[i]
                elif str(data.dtype.names[j]) == "PRI":
                    cleaned_data[j] = pri[i]
                else:
                    continue

            cleaned.writerow(cleaned_data)

            # Give user an idea how much has been completed.
            if i % 5000 == 0:
                print(".")
            if i % 10000 == 0:
                print(".. Row " + str(i))

    print("Data has been cleaned!\n*\n**\n***")
    del data, cire, mtci, ccci, cccia, ci800, datt, datta
    del ndare, ndre, ndrre, ndvia, ndvir, nvg2, nvg800, pri
    return


def Clean_IRT_Plot(clean_csv, forward_file, nadir_file, zeros_toggle,
                   custom_range, range_onoff):
    """This function  reads in the csv from user and returns cleaned data

    We assign corresponding data to columns. Then adjust the data to fit within
    a certain range. if out of bounds, reassign a nan value. Then write clean
    data to new csv.
    clean_csv -> str
    forward_file -> str
    nadir_file -> str
    zeros_toggle -> int
    custon_range -> list of floats
    range_onoff -> int
    """

    # This often gets run through twice.
    # The initial cleaning the file names are different
    # The preened cleaning will have two/three of the same file names to
    # execute the preening on the already cleaned file
    if clean_csv != forward_file:
        print("***\n**\n*\nLoading in the raw data from " +
              forward_file + "\n.\n.")

        dataf = np.genfromtxt(forward_file, dtype=None, delimiter=",",
                              names=True, skip_header=0)

        print("Loading in the raw data from " + nadir_file + "\n.\n.")
        datan = np.genfromtxt(nadir_file, dtype=None, delimiter=",",
                              names=True, skip_header=0)

        target_cel_forward = dataf['Target_Cel'].astype(np.float)
        target_cel_nadir = datan['Target_Cel'].astype(np.float)

        target_cel = [target_cel_forward, target_cel_nadir]
    else:
        print("***\n**\n*\nLoading in the data from\n" +
              clean_csv + " to submit preening\n.\n.")

        data = np.genfromtxt(clean_csv, dtype=None, delimiter=",",
                             names=True, skip_header=0)

        target_cel = data['Target_Cel']

        # this is to address a mysterious str issue
        for element in range(len(target_cel)):
            if isinstance(target_cel[element], str):
                target_cel[element] = np.nan

        target_cel = target_cel.astype(np.float)

        target_cel = [target_cel]

    # some lines have all nan values and it
    # throws a warning about Runtime.
    # Nothing to worry about, just numpy at the edges
    # of its formal structure
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", category=RuntimeWarning)

        # this means the range_toggle is off and we should include the range.
        if range_onoff == 0:

            c_min = custom_range[0]
            c_max = custom_range[1]
            print("Removing values outside " + str(c_min) +
                  " and " + str(c_max))

            # data outside of range gets nan value
            for m in range(len(target_cel)):
                target_cel[m][(target_cel[m] < c_min) |
                              (target_cel[m] > c_max)] = np.nan

        # If zero toggle is on then replace zeros with nan
        if zeros_toggle == 1:

            print("Zeros being eliminated")
            # for elements in array,
            # replace values outside of the range with nan
            for n in range(len(target_cel)):
                target_cel[n][(target_cel[n] == 0.)] = np.nan

    # write untouched val and modified array to new file
    with open(clean_csv, 'wb') as cc:
        cleaned = csv.writer(cc, delimiter=",")

        for k in range(len(target_cel)):
            if k == 0 and (clean_csv != forward_file):
                data = dataf
            if k == 1 and (clean_csv != forward_file):
                data = datan

            # i marks the current row in spreadsheet
            for i in range(len(target_cel[k])):

                if i == 0:
                    cleaned.writerow(data.dtype.names)

                cleaned_data = data[i]

                # j marks element in each row
                for j in range(len(data.dtype.names)):

                    # replace out of bounds data with nan
                    if str(data.dtype.names[j]) == "Target_Cel":
                        cleaned_data[j] = target_cel[k][i]
                    else:
                        continue

                cleaned.writerow(cleaned_data)

                # Give user an idea how much has been completed.
                if i % 5000 == 0:
                    print(".")
                if i % 10000 == 0:
                    print(".. Row " + str(i))

    print("Data has been cleaned!\n*\n**\n***")
    del data, target_cel
    return


def Clean_HonPep_Plot(clean_csv, honpep_file, zeros_toggle,
                      custom_range, range_onoff):
    """This function  reads in the csv from user and returns cleaned data

    We assign corresponding data to columns. Then adjust the data to fit within
    a certain range. if out of bounds, reassign a nan value. Then write clean
    data to new csv.
    clean_csv -> str
    honpep_file -> str
    zeros_toggle -> int
    custon_range -> list of floats
    range_onoff -> int
    """

    print("***\n**\n*\nLoading in the raw data from " +
          honpep_file + "\n.\n.")

    data = np.genfromtxt(honpep_file, dtype=None, delimiter=",",
                         names=True, skip_header=0)

    crop_height = data['CropHeight'].astype(np.float)

    # some lines have all nan values and it
    # throws a warning about Runtime.
    # Nothing to worry about, just numpy at the edges
    # of its formal structure
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", category=RuntimeWarning)

        # this means the range_toggle is off and we should include the range.
        if range_onoff == 0:

            c_min = custom_range[0]
            c_max = custom_range[1]
            print("Removing values outside " + str(c_min) +
                  " and " + str(c_max))

            # data outside of range gets nan value
            crop_height[(crop_height < c_min) |
                        (crop_height > c_max)] = np.nan

        # If zero toggle is on then replace zeros with nan
        if zeros_toggle == 1:

            print("Zeros being eliminated")
            # for elements in array, replace zeros with nan
            crop_height[(crop_height == 0.)] = np.nan

    # write untouched val and modified array to new file
    with open(clean_csv, 'wb') as cc:
        cleaned = csv.writer(cc, delimiter=",")

        # i marks the current row in spreadsheet
        for i in range(len(crop_height)):

            if i == 0:
                cleaned.writerow(data.dtype.names)

            cleaned_data = data[i]

            # j marks element in each row
            for j in range(len(data.dtype.names)):

                # replace out of bounds data with nan
                if str(data.dtype.names[j]) == "CropHeight":
                    cleaned_data[j] = crop_height[i]
                else:
                    continue

            cleaned.writerow(cleaned_data)

            # Give user an idea how much has been completed.
            if i % 5000 == 0:
                print(".")
            if i % 10000 == 0:
                print(".. Row " + str(i))

    print("Data has been cleaned!\n*\n**\n***")
    del data, crop_height
    return


def Create_Preen_File(housing_folder, clean_csv):
    """This will create a file for the bad sensors.

    This gets called one of two ways, through the UI toggle or through Stat An
    housing_folder -> str
    clean_csv -> str
    """

    preen_header = ["Trait", "Plot", "Sensor"]

    csv_path, csv_name = os.path.split(clean_csv)

    split_csv = csv_name.split("_")

    preen_name = "_".join(split_csv[:-1]) + "_preened.csv"

    preen_file = housing_folder + preen_name

    with open(preen_file, 'wb') as pf:
        pf_write = csv.writer(pf, delimiter=",")

        pf_write.writerow(preen_header)

    return


def Append_Preen_File(housing_folder, clean_csv, bad_sensors):
    """Writes the bad sensors to the preen file

    housing_folder -> str
    clean_csv -> str
    bad_sensors -> list of strs
    """

    csv_path, csv_name = os.path.split(clean_csv)

    split_csv = csv_name.split("_")

    preen_name = "_".join(split_csv[:-1]) + "_preened.csv"

    preen_file = housing_folder + preen_name

    with open(preen_file, 'ab') as pf:
        pf_write = csv.writer(pf, delimiter=",")

        pf_write.writerow(bad_sensors)

    return


def Preen_CC_Plot(clean_csv, preen_csv):
    """Analyzes the sesnors that outside the given plot range.

    If the sensors are outside, then they get written to the preen file.
    clean_csv -> str
    preen_csv -> str
    """

    with open(preen_csv, 'rb') as pcsv:
        preen_rows = list(csv.reader(pcsv, delimiter=','))

    if len(preen_rows) == 1:
        # if we are here, the file is empty so exit out
        return

    print("***\n**\n*\nInitiating Preening\n.\n.\nEvaluating Bad Sensors...")

    tps = []

    for a in range(1, len(preen_rows)):

        # incase there's been a misprinting in the file
        if len(preen_rows[a]) != 3:
            continue

        tps.append((preen_rows[a][0], preen_rows[a][1], preen_rows[a][2]))

    print(".\n.\nTotal Sensors To Be Culled: " + str(len(tps)))

    print(".\n.\nCurrently Culling...\n.")

    # Find a unique table name
    csv_path, csv_name = os.path.split(preen_csv)
    p_table = csv_name.split(".")
    preen_table = p_table[0]

    df = pandas.read_csv(clean_csv, low_memory=False)
    temp_preen_db = "Other_Temp_Enchilada.db"
    if os.path.exists(temp_preen_db) is True:
        os.remove(temp_preen_db)
    con = sqlite3.connect(temp_preen_db)
    df.to_sql(preen_table, con)

    # A new one will be written later with the same name
    os.remove(clean_csv)

    with con:

        cur = con.cursor()

        # This is susceptable to SQL injection attacks
        for b in range(len(tps)):
            trait = tps[b][0]
            plot = tps[b][1]
            sensor = tps[b][2]

            cur.execute('UPDATE ' + preen_table + ' ' +
                        'SET ' + trait + '="nan" '
                        'WHERE PLOT=? '
                        'AND Sensor=?', (plot, sensor))

        con.commit()
        cur.execute('SELECT * FROM ' + preen_table)

        table = pandas.read_sql('SELECT * FROM ' + preen_table,
                                con).drop(['index'], axis=1)

        table.to_csv(clean_csv, index=False)

    print("Data Preened!\n*\n**\n***")

    del cur
    del con
    del df
    del table

    os.remove(temp_preen_db)

    return


def Find_Run(housing_folder):
    """Identify which run is being executed via the file path

    housing folder -> str
    """

    folder_split = housing_folder.split("_")

    run = ""

    for fs in folder_split:
        if fs.startswith("Run"):
            run = "_" + fs

    if len(run) == 0:
        print("Oh no! Program cannot identify which Run " +
              housing_folder + " belongs too! Please be sure"
              " to have '_Run##_' somewhere in the filepath")
        sys.exit(1)

    return run


def Stat_Analysis(housing_folder, clean_csv, zeros_toggle,
                  reanalyze_toggle, plot_std_toggle,
                  plot_std, range_onoff, custom_range,
                  stressors, species, graph_toggle):
    """This is to give a bit of statistical analysis on data

    We are looking to find the mean, stdev, and range of each column
    So we create a new file with those associated values.
    And then we make graphs so a user so can visually validate the ranges
    and to kick out the outliers
    housing_folder -> str
    clean_csv -> str
    zeros_toggle -> int
    reanalyze_toggle -> int
    plot_std_toggle -> int
    plot_std -> int
    range_onoff -> int
    custom_range -> list of floats
    num_exp ->
    stressors -> list of str
    species -> list of str
    graph_toggle -> int
    """

    print("***\n**\n*\nBeginning Stat Analysis\n.\n.")

    csv_path, csv_name = os.path.split(clean_csv)
    table = csv_name.split(".")
    clean_table = table[0]

    # have pandas read in clean_csv then convert to db
    # for a little SQLite functionality
    df = pandas.read_csv(clean_csv, low_memory=False, error_bad_lines=False)
    temp_db = housing_folder + "\Temp_Enchilada.db"
    if os.path.exists(temp_db) is True:
        os.remove(temp_db)
    con = sqlite3.connect(temp_db)
    df.to_sql(clean_table, con)

    Create_Preen_File(housing_folder, clean_csv)

    # ID which run via we are in via the file name
    run = Find_Run(housing_folder)

    traits_run = []
    # use db searching and then perform stat analysis
    with con:

        cur = con.cursor()

        csv_path, csv_name = os.path.split(clean_csv)
        csv_split = csv_name.split("_")
        tool = csv_split[0]

        if csv_name == "CC_plots_cleaned.csv":
            cc_headers = ["CIRE", "MTCI", "CCCI", "CCCIA",
                          "CI800", "DATT", "DATTA", "NDARE",
                          "NDRE", "NDRRE", "NDVIA", "NDVIR",
                          "NVG2", "NVG800", "PRI"]

            for i in range(len(cc_headers)):

                traits_run.append(cc_headers[i] + "_CC" + run)

        elif csv_name == "IRT_cleaned.csv":
            cc_headers = ["Target_Cel"]
            traits_run = [cc_headers[0] + "_IRT" + run]

        elif csv_name == "Honey_plots_cleaned.csv":
            cc_headers = ["CropHeight"]
            traits_run = [cc_headers[0] + "_Honey" + run]
        elif csv_name == "Pep_plots_cleaned.csv":
            cc_headers = ["CropHeight"]
            traits_run = [cc_headers[0] + "_Pep" + run]

        # query database for columns
        cur.execute("SELECT FID, PLOT " +
                    "FROM " + clean_table)

        # put results into a variable we can search through
        rows = cur.fetchall()

        all_plots = []

        # find all plot names, without repeats
        for i in range(len(rows)):
            plot_name = str(rows[i][1])

            if plot_name not in all_plots:
                all_plots.append(plot_name)

        # Alphabetize in reverse
        all_plots.sort(reverse=True)

        # Define list of lists where the sublist is for a single species
        experiments = [[] for i in range(len(species))]

        # group of for loops to put each plot in a corresponding experiment
        # This will create a list and group the plots by Species
        for i in range(len(species)):
            for j in range(len(stressors)):
                for k in range(len(all_plots)):
                    plot_split = list(all_plots[k])

                    # string matching to place each plot with like plots
                    if plot_split[0] == stressors[j] and \
                       plot_split[-1] == species[i]:
                        experiments[i].append(all_plots[k])

    del con

    if csv_name == "CC_plots_cleaned.csv":

        # packages data into a tuple for the arguments for multprocessing
        cire_args = ("CIRE", housing_folder, tool, traits_run[0],
                     species, plot_std_toggle, experiments,
                     clean_table, plot_std, clean_csv,
                     temp_db, range_onoff, custom_range, graph_toggle)
        mtci_args = ("MTCI", housing_folder, tool, traits_run[1],
                     species, plot_std_toggle, experiments,
                     clean_table, plot_std, clean_csv,
                     temp_db, range_onoff, custom_range, graph_toggle)
        ccci_args = ("CCCI", housing_folder, tool, traits_run[2],
                     species, plot_std_toggle, experiments,
                     clean_table, plot_std, clean_csv,
                     temp_db, range_onoff, custom_range, graph_toggle)
        cccia_args = ("CCCIA", housing_folder, tool, traits_run[3],
                      species, plot_std_toggle, experiments,
                      clean_table, plot_std, clean_csv,
                      temp_db, range_onoff, custom_range, graph_toggle)
        ci800_args = ("CI800", housing_folder, tool, traits_run[4],
                      species, plot_std_toggle, experiments,
                      clean_table, plot_std, clean_csv,
                      temp_db, range_onoff, custom_range, graph_toggle)
        datt_args = ("DATT", housing_folder, tool, traits_run[5],
                     species, plot_std_toggle, experiments,
                     clean_table, plot_std, clean_csv,
                     temp_db, range_onoff, custom_range, graph_toggle)
        datta_args = ("DATTA", housing_folder, tool, traits_run[6],
                      species, plot_std_toggle, experiments,
                      clean_table, plot_std, clean_csv,
                      temp_db, range_onoff, custom_range, graph_toggle)
        ndare_args = ("NDARE", housing_folder, tool, traits_run[7],
                      species, plot_std_toggle, experiments,
                      clean_table, plot_std, clean_csv,
                      temp_db, range_onoff, custom_range, graph_toggle)
        ndre_args = ("NDRE", housing_folder, tool, traits_run[8],
                     species, plot_std_toggle, experiments,
                     clean_table, plot_std, clean_csv,
                     temp_db, range_onoff, custom_range, graph_toggle)
        ndrre_args = ("NDRRE", housing_folder, tool, traits_run[9],
                      species, plot_std_toggle, experiments,
                      clean_table, plot_std, clean_csv,
                      temp_db, range_onoff, custom_range, graph_toggle)
        ndvia_args = ("NDVIA", housing_folder, tool, traits_run[10],
                      species, plot_std_toggle, experiments,
                      clean_table, plot_std, clean_csv,
                      temp_db, range_onoff, custom_range, graph_toggle)
        ndvir_args = ("NDVIR", housing_folder, tool, traits_run[11],
                      species, plot_std_toggle, experiments,
                      clean_table, plot_std, clean_csv,
                      temp_db, range_onoff, custom_range, graph_toggle)
        nvg2_args = ("NVG2", housing_folder, tool, traits_run[12],
                     species, plot_std_toggle, experiments,
                     clean_table, plot_std, clean_csv,
                     temp_db, range_onoff, custom_range, graph_toggle)
        nvg800_args = ("NVG800", housing_folder, tool, traits_run[13],
                       species, plot_std_toggle, experiments,
                       clean_table, plot_std, clean_csv,
                       temp_db, range_onoff, custom_range, graph_toggle)
        pri_args = ("PRI", housing_folder, tool, traits_run[14],
                    species, plot_std_toggle, experiments,
                    clean_table, plot_std, clean_csv,
                    temp_db, range_onoff, custom_range, graph_toggle)

        trait_args = [cire_args, mtci_args, ccci_args, cccia_args,
                      ci800_args, datt_args, datta_args, ndare_args,
                      ndre_args, ndrre_args, ndvia_args, ndvir_args,
                      nvg2_args, nvg800_args, pri_args]

        # launch multiprocessing, utilizing 5 threads
        start_time = time.time()
        pool = Pool(5)
        pool.map(Stat_Core, trait_args)

        print("--- %s seconds ---" % round((time.time() - start_time), 2))
    else:

        trait = cc_headers[0]
        core_args = (trait, housing_folder, tool, traits_run[0], species,
                     plot_std_toggle, experiments, clean_table,
                     plot_std, clean_csv, temp_db, range_onoff,
                     custom_range, graph_toggle)

        start_time = time.time()
        Stat_Core(core_args)
        print("--- %s seconds ---" % round((time.time() - start_time), 2))

    del cur
    del df
    del rows

    os.remove(temp_db)

    return


def Stat_Core(core_args):
    """Execute the stat analysis

    core_args -> mixed list
    """

    trait = core_args[0]
    housing_folder = core_args[1]
    tool = core_args[2]
    traits_run = core_args[3]
    species = core_args[4]
    plot_std_toggle = core_args[5]
    experiments = core_args[6]
    clean_table = core_args[7]
    plot_std = core_args[8]
    clean_csv = core_args[9]
    temp_db = core_args[10]
    range_onoff = core_args[11]
    custom_range = core_args[12]
    graph_toggle = core_args[13]

    trait_avgs = {}

    con = sqlite3.connect(temp_db)

    # The meat of SA. If allowed, finds stdev and filters out some sensors
    # Also plots the data being calculated
    with con:

        cur = con.cursor()

        print("Analyzing Trait " + trait)

        column_data = {}
        trait_avgs = []

        # Define stat file name
        experiment_file = housing_folder + "Stats_" + \
            str(tool) + "_" + str(trait) + ".csv"

        exp_header = ["Plot", traits_run]

        with open(experiment_file, 'wb') as exp_file:
            write_header = csv.writer(exp_file, delimiter=',')
            write_header.writerow(exp_header)

            new_stat_file = csv.writer(exp_file, delimiter=',')

            # Define pdf file the plotted sensors
            sen_dat_file = housing_folder + "\\" + tool + "_" + trait + ".pdf"

            if graph_toggle == 1:
                # Enables multiples pages
                pp = PdfPages(sen_dat_file)

            # Loop over the amount of different species avaialable
            for d in range(len(species)):

                if plot_std_toggle == 1:
                    std_dict = {}

                all_sensors = {}

                # Loop over all plots with a group of species
                for j in range(len(experiments[d])):

                    current_plot = experiments[d][j]
                    print trait + " " + current_plot

                    # find all sensors with matching plot id
                    cur.execute("SELECT PLOT, Sensor "
                                "FROM " + clean_table + " " +
                                "WHERE PLOT='" + current_plot + "'")

                    plot_data = cur.fetchall()

                    plot_sensors = []

                    # Loop over sensors available
                    for a in range(len(plot_data)):
                        sensor_id = str(plot_data[a][1])

                        if sensor_id not in plot_sensors:
                            # Save names of avaialable sensors
                            plot_sensors.append(sensor_id)

                    # Because I can
                    plot_sensors.sort(reverse=True)

                    # Associate the available sensors to a pspecific plot
                    all_sensors.update({current_plot: plot_sensors})

                    all_sensor_data = []

                    plot_totals = np.array([])

                    # Collect data from each sensor
                    for b in range(len(plot_sensors)):

                        # plot_std_toggle analysis is here
                        if b == 0 and plot_std_toggle == 1:

                            cur.execute("SELECT " + trait + " "
                                        "FROM " + clean_table + " " +
                                        "WHERE PLOT='" + current_plot + "'")

                            plot_data = np.array(cur.fetchall(),
                                                 dtype=np.float).T

                            plot_data[0] = plot_data[0].astype(np.float)

                            # Calc the Stand Dev and mean of all data
                            plot_mean_1 = float(np.nanmean(plot_data[0]))
                            plot_stdev_1 = float(np.nanstd(plot_data[0]))

                            # Define a pseudo range within +/- std of mean
                            plot_min_1 = plot_mean_1 - plot_stdev_1
                            plot_max_1 = plot_mean_1 + plot_stdev_1

                            with warnings.catch_warnings():
                                warnings.simplefilter(
                                    "ignore",
                                    category=RuntimeWarning)
                                # Find all data within pseudo range
                                plot_1 = plot_data[0][(plot_data[0] >
                                                       plot_min_1) &
                                                      (plot_data[0] <
                                                       plot_max_1)]

                            # Take mean and Stand Dev of
                            # data with in psuedo range
                            plot_mean = float(np.nanmean(plot_1))
                            plot_stdev = float(np.nanstd(plot_1))

                            # Define new pseudo range on new stdev
                            plot_min = plot_mean - plot_stdev * plot_std
                            plot_max = plot_mean + plot_stdev * plot_std

                            # Associate plot with calc'd data
                            std_dict.update({current_plot: [plot_min,
                                                            plot_max,
                                                            plot_mean]})

                        current_sensor = plot_sensors[b]

                        cur.execute("SELECT " + trait + " " +
                                    "FROM " + clean_table + " " +
                                    "WHERE (PLOT='" + current_plot + "') " +
                                    "AND (Sensor='" + current_sensor + "') ")

                        plot_sensor = np.array(cur.fetchall(),
                                               dtype=np.float)

                        stat_data_t = np.transpose(plot_sensor)

                        # some lines have all nan values and it
                        # throws a warning about Runtime.
                        # Nothing to worry about, just numpy at the edges
                        # of its formal structure
                        with warnings.catch_warnings():
                            warnings.simplefilter("ignore",
                                                  category=RuntimeWarning)

                            # Calc the Mean
                            mean = float(np.nanmean(stat_data_t[0]))

                            # Filter the sensors
                            if plot_std_toggle == 1 and \
                                    (mean < plot_min or mean > plot_max):

                                stat_data_t[0] = np.nan

                                bad_sensors = [trait, current_plot,
                                               current_sensor]

                                # Send info to preen file
                                Append_Preen_File(housing_folder,
                                                  clean_csv,
                                                  bad_sensors)

                        if plot_std_toggle == 0:
                            plot_totals = np.append(plot_totals,
                                                    stat_data_t[0])
                        elif plot_std_toggle == 1:
                            if mean > plot_min and mean < plot_max:
                                plot_totals = np.append(plot_totals,
                                                        stat_data_t[0])

                        sensor_column = stat_data_t[0].astype(np.float)

                        # Add data to collection for the plot
                        all_sensor_data.append(sensor_column)

                    # Replace some data
                    for x in range(len(all_sensor_data)):
                        all_sensor_data[x] = all_sensor_data[x]\
                                           [~np.isnan(all_sensor_data[x])]

                    # Associate plot with sensor data
                    column_data.update({current_plot: all_sensor_data})

                    mean_plot = float(np.nanmean(plot_totals))

                    all_stats = [current_plot, mean_plot]
                    trait_avgs.append(all_stats)

                if plot_std_toggle == 0:
                    std_dict = 0

                if graph_toggle == 1:
                    Plot_The_Data(housing_folder, experiments[d], column_data,
                                  all_sensors, std_dict, trait,
                                  plot_std_toggle, range_onoff,
                                  custom_range, clean_csv, pp)

            for a in range(len(trait_avgs)):
                new_stat_file.writerow(trait_avgs[a])

            if graph_toggle == 1:
                pp.close()
                del pp
    del con
    del cur
    return


def Find_Matrix(len_exp):
    """"Define the most square-lke matrix in which to display data.

    len_exp -> int
    """

    prime_num = 0

    # Loop over ints until len_exp is between two neighboring squares
    while True:

        next_prime = prime_num + 1

        square_num = prime_num ** 2
        next_square = next_prime ** 2

        if len_exp >= square_num and len_exp < next_square:

            if len_exp == square_num:
                rows = prime_num
                columns = prime_num
                dummy = 0
            elif len_exp <= prime_num * next_prime:
                rows = next_prime
                columns = prime_num
                dummy = next_prime * prime_num - len_exp
            else:
                rows = next_prime
                columns = next_prime
                dummy = next_square - len_exp

            break

        prime_num += 1

    return rows, columns, dummy


def Plot_The_Data(housing_folder, experiment, column_data,
                  all_sensors, std_dict, trait, plot_std_toggle,
                  range_onoff, custom_range, clean_csv, pp):
    """To Plot all of the beautiful data we have calculated

    housing_folder -> str
    experiment -> list of str (plots within experiment)
    column_data -> dict with plot as key and list of array of sensor points
    all_sensors -> dict of plot as key and list of sensors as value
    std_dict -> dict of plot as key and list of mean, and +\- std
    trait -> str
    plot_std_toggle -> int
    range_onoff -> int
    custom_range -> list
    clean_csv -> str
    pp -> PDFPages object to include multiple images in a file
    """

    csv_path, csv_name = os.path.split(clean_csv)

    # Find rows x cols for a square matrix plots can fit into
    n_rows, n_cols, dummy = Find_Matrix(len(experiment))

    # add dummy vars so we can make a square subplot safely
    dummy_list = ['dummy'] * dummy
    experiment = experiment + dummy_list

    print("Creating image for " + trait)
    expr = np.reshape(np.array(experiment), (n_rows, n_cols))

    fig, axes = plt.subplots(nrows=n_rows, ncols=n_cols,
                             sharey=True, figsize=(9*n_cols, 5*n_rows))

    # Loop for creating subplots for each image
    for a in range(n_rows):
        for b in range(n_cols):

            # If dummy variable, just pass
            try:
                all_sensors[expr[a][b]]
                column_data[expr[a][b]]
            except KeyError:
                continue

            tot_sensors = len(all_sensors[expr[a][b]]) + 1
            x_sensors = range(tot_sensors)

            if plot_std_toggle == 1:
                std_bounds = std_dict[expr[a][b]]
                std_min = std_bounds[0]
                std_max = std_bounds[1]
                std_mean = std_bounds[2]

                axes[a, b].plot(x_sensors, [std_min]*tot_sensors,
                                c='red', ls='--')
                axes[a, b].plot(x_sensors, [std_max]*tot_sensors,
                                c='red', ls='--')
                axes[a, b].plot(x_sensors, [std_mean]*tot_sensors,
                                c='pink')

            # If Memory is an issue (too big small) then avoid making plot
            try:
                axes[a, b].boxplot(column_data[expr[a][b]],
                                   labels=all_sensors[expr[a][b]])
            except MemoryError:
                print("Failed to make boxplot for " + expr[a][b])
                continue

            axes[a, b].set_title(expr[a][b], fontsize=10)

            if range_onoff == 0:
                axes[a, b].set_ylim([custom_range[0], custom_range[1]])

    plt.suptitle(trait, fontsize=20)

    print("-- Made Image for trait " + trait)

    try:
        pp.savefig()
    except MemoryError:
        if expr[a][b] != "dummy":
            print("Failed to save plot for trait " + expr[a][b])
        pass

    fig.clf()
    plt.close(fig)

    del fig
    del axes

    return


def Combine_Stats(top_folder):
    """Combine the Stat_.csv files into one xlsx file

    top_folder -> str
    """

    match_list = []
    stat_file = top_folder + "\Complete_Stats.xlsx"

    for root, dirnames, filenames in os.walk(top_folder):
        for filename in fnmatch.filter(filenames, 'Stats_*.csv'):
            match_list.append(os.path.join(root, filename))

    all_stats = {}
    plot_amount = 0

    # find all of the plots avaialble. some files are fising a few plots
    # if the file with the most amount of plots is missing a few, then trouble
    for f in range(len(match_list)):

        data = np.genfromtxt(match_list[f], dtype=None, delimiter=",",
                             names=True, skip_header=0)

        if len(data[data.dtype.names[0]]) > plot_amount:

            plot_ids = tuple(data[data.dtype.names[0]])
            all_stats.update({"PLOTS": plot_ids})

            plot_amount = len(plot_ids)

    # Associate the trait averages to the run-trait-sensor
    for g in range(len(match_list)):

        data = np.genfromtxt(match_list[g], dtype=None, delimiter=",",
                             names=True, skip_header=0)

        plot_check = {}

        trait_avgs = [np.nan] * len(plot_ids)

        for row in range(len(data)):

            plot_check.update({data[row][0]: data[row][1]})

        for h in range(len(plot_ids)):

            pid = plot_ids[h]

            if pid in plot_check.keys():

                trait_avgs[h] = plot_check[pid]

        all_stats.update({data.dtype.names[1]: trait_avgs})

    key_headers = all_stats.keys()
    key_headers.sort()
    key_headers.remove("PLOTS")

    key_headers_split = []

    book = Workbook()
    ws = book.active

    bad_sheet = book.get_sheet_by_name('Sheet')
    book.remove_sheet(bad_sheet)

    for k in range(len(key_headers)):
        key_headers_split.append(key_headers[k].split("_"))

    for key, group in itertools.groupby(key_headers_split, lambda x: x[0:2]):

        sheet = book.create_sheet(title="_".join(key))
        all_runs = ["PLOTS"]

        for thing in group:

            all_runs.append("_".join(thing))

        for r in range(len(plot_ids)):

            combined = []

            for col in range(len(all_runs)):

                if col == 0:
                    combined.append(plot_ids)
                else:
                    combined.append(all_stats[all_runs[col]])

            if r == 0:

                for h in range(len(all_runs)):
                    sheet.cell(column=h+1, row=r+1, value=str(all_runs[h]))

            for c in range(len(combined)):
                sheet.cell(column=c+1, row=r+2, value=combined[c][r])

    book.save(filename=stat_file)

    print "Finished merging!\n*\n**\n***"
    return


def Clean_Up(top_folder):

    for root, dirnames, filenames in os.walk(top_folder):
        for filename in fnmatch.filter(filenames, 'Stats_*.csv'):
            os.remove(filename)

    return


def File_Check(match_folder, cc_toggle, irt_toggle, honey_toggle, pep_toggle):
    """Provide a preliminary check that every file is available to later be ran

    match_folder -> str
    cc_toggle -> int
    irt_toggle -> int
    pep_toggle -> int
    honey_toggle -> int
    """

    missing_files = []

    for housing_folder in match_folder:
        if cc_toggle == 1:
            cc_path = housing_folder + "\CC_plots"
            if os.path.exists(cc_path + ".csv") is False and \
               os.path.exists(cc_path + ".txt") is False:
                missing_files.append(cc_path + ".csv")

        if irt_toggle == 1:
            irtf_path = housing_folder + "\IRT_forward_plots"
            irtn_path = housing_folder + "\IRT_nadir_plots"
            if os.path.exists(irtf_path + ".csv") is False and \
               os.path.exists(irtf_path + ".txt") is False:
                missing_files.append(irtf_path + ".csv")
            if os.path.exists(irtn_path + ".csv") is False and \
               os.path.exists(irtn_path + ".txt") is False:
                missing_files.append(irtn_path + ".csv")

        if pep_toggle == 1:
            pep_path = housing_folder + "\Pep_plots"
            if os.path.exists(pep_path + ".csv") is False and \
               os.path.exists(pep_path + ".txt") is False:
                missing_files.append(pep_path + ".csv")

        if honey_toggle == 1:
            honey_path = housing_folder + "\Honey_plots"
            if os.path.exists(honey_path + ".csv") is False and \
               os.path.exists(honey_path + ".txt") is False:
                missing_files.append(honey_path + ".csv")

    if len(missing_files) > 0:
        print(" ...---... " * 5 + "\n")
        print("Oh no! There are missing files!")

        for file_ in missing_files:
            print file_

        raw_input(" ...---... " * 5)
        sys.exit(1)

    return


def Parameters_Executed(housing_folder, inputs):
    """Provide a history of what was executed

    housing_folder -> str
    inputs -> list of str
    """

    parameters = []

    now = datetime.datetime.now()
    now_time = str(now.hour) + ":" + str(now.minute)
    now_date = str(now.month) + "/" + str(now.day) + "/" + str(now.year)

    parameters.append("##\n## A Catalog of the Paramters"
                      "## used in the lastest execution\n"
                      "## of The_Cleaners.py\n##")
    parameters.append(" ")
    parameters.append("## If 'Executed' appears in a line, the associated\n"
                      "## number is a Binary representation of True/False")
    parameters.append(" ")
    parameters.append("Time: " + now_time + "    Date: " + now_date)
    parameters.append(" ")

    parameters.append("Top Folder: " + inputs[0])
    parameters.append("Stressors: " + inputs[1])
    parameters.append("Species : " + inputs[2])
    parameters.append(" ")

    parameters.append("CC Plots Executed: " + inputs[3])
    parameters.append("Custom Range: " + inputs[4])
    parameters.append("Do Not Define Range Executed: " + inputs[5])
    parameters.append("Plot Standard Deviation: " + inputs[6])
    parameters.append("Plot Standard Deviation Executed: " + inputs[7])
    parameters.append("Remove Zeros Executed: " + inputs[8])
    parameters.append("Graphs Executed: " + inputs[9])
    parameters.append("Reanalyzed Executed: " + inputs[10])
    parameters.append(" ")

    parameters.append("IRT Plots Executed: " + inputs[11])
    parameters.append("Custom Range: " + inputs[12])
    parameters.append("Do Not Define Range Executed: " + inputs[13])
    parameters.append("Plot Standard Deviation: " + inputs[14])
    parameters.append("Plot Standard Deviation Executed: " + inputs[15])
    parameters.append("Remove Zeros Executed: " + inputs[16])
    parameters.append("Graphs Executed: " + inputs[17])
    parameters.append("Reanalyzed Executed: " + inputs[18])
    parameters.append(" ")

    parameters.append("Honeywell Plots Executed: " + inputs[19])
    parameters.append("Custom Range: " + inputs[20])
    parameters.append("Do Not Define Range Executed: " + inputs[21])
    parameters.append("Plot Standard Deviation: " + inputs[22])
    parameters.append("Plot Standard Deviation Executed: " + inputs[23])
    parameters.append("Remove Zeros Executed: " + inputs[24])
    parameters.append("Graphs Executed: " + inputs[25])
    parameters.append("Reanalyzed Executed: " + inputs[26])
    parameters.append(" ")

    parameters.append("Pepperl Plots Executed: " + inputs[27])
    parameters.append("Custom Range: " + inputs[28])
    parameters.append("Do Not Define Range Executed: " + inputs[29])
    parameters.append("Plot Standard Deviation: " + inputs[30])
    parameters.append("Plot Standard Deviation Executed: " + inputs[31])
    parameters.append("Remove Zeros Executed: " + inputs[32])
    parameters.append("Graphs Executed: " + inputs[33])
    parameters.append("Reanalyzed Executed: " + inputs[34])

    Create_And_Write_File(housing_folder + "Parameters_Used", parameters)

    return


###############################################################################
##~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ Main Function ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~##
###############################################################################
def main():

    # Launch the GUI
    User_Interface()

    inputs = Read_Text_File("USE_Cleaner.txt")

    np.set_printoptions(threshold=np.nan)

    top_folder = inputs[0]
    stressors = inputs[1]
    species = inputs[2]

    cc_toggle = int(inputs[3])
    custom_range_c = inputs[4]
    range_toggle_c = int(inputs[5])
    plot_std_c = inputs[6]
    plot_std_toggle_c = int(inputs[7])
    zeros_toggle_c = int(inputs[8])
    graph_toggle_c = int(inputs[9])
    reanalyze_toggle_c = int(inputs[10])

    irt_toggle = int(inputs[11])
    custom_range_i = inputs[12]
    range_toggle_i = int(inputs[13])
    plot_std_i = inputs[14]
    plot_std_toggle_i = int(inputs[15])
    zeros_toggle_i = int(inputs[16])
    graph_toggle_i = int(inputs[17])
    reanalyze_toggle_i = int(inputs[18])

    honey_toggle = int(inputs[19])
    custom_range_h = inputs[20]
    range_toggle_h = int(inputs[21])
    plot_std_h = inputs[22]
    plot_std_toggle_h = int(inputs[23])
    zeros_toggle_h = int(inputs[24])
    graph_toggle_h = int(inputs[25])
    reanalyze_toggle_h = int(inputs[26])

    pep_toggle = int(inputs[27])
    custom_range_p = inputs[28]
    range_toggle_p = int(inputs[29])
    plot_std_p = inputs[30]
    plot_std_toggle_p = int(inputs[31])
    zeros_toggle_p = int(inputs[32])
    graph_toggle_p = int(inputs[33])
    reanalyze_toggle_p = int(inputs[34])

    match_folders = []

    for root, dirnames, filenames in os.walk(top_folder):

        possible_files = ["CC_plots.csv", "IRT_forward_plots.csv",
                          "IRT_nadir_plots.csv", "Honey_plots.csv",
                          "Pep_plots.csv", "cc_plots.csv",
                          "irt_forward_plots.csv", "irt_nadir_plots.csv",
                          "honey_plots.csv", "pep_plots.csv", "CC_plots.txt",
                          "IRT_forward_plots.txt", "IRT_nadir_plots.txt",
                          "Honey_plots.txt", "Pep_plots.txt"]

        # if any files match in pf and filenames and root is not in match_list
        if any(i for i in possible_files if i in filenames) and \
           root not in match_folders and "re-filtered" not in root:

            match_folders.append(root)

    #match_folders.sort(reverse=True)
    print "--------\nFolders found:"
    for file_ in match_folders:
        print file_
    print "--------\n*\n**\n***"

    oops = ".\n.\nOops! "
    range_exep = " Range was not a valid number. " + \
                 "Be sure to use numbers and don't incorporate spaces\n" + \
                 "For example:  0.2,0.9345\n.\n.\n" + \
                 "Quiting Program..."

    plot_exep = " Plot Standard Deviation Scalar " + \
                "was not a valid number.\n" + \
                "Be sure to include real numbers greater than 0\n" + \
                ".\n.\nQuiting Program..."

    # Validate all of the  the necessary user inputs for correct type
    if os.path.exists(top_folder) is False:

        os.path.exists(top_folder)
        raw_input(top_folder + " does not exist!\n"
                  "Try again with an existing file or directory\n"
                  "Quiting Program...")
        sys.exit(1)

    stressors = stressors.split(",")
    species = species.split(",")

    if range_toggle_c == 0:

        custom_range_c = custom_range_c.split(",")

        try:
            custom_range_c = [
                           float(custom_range_c[0]),
                           float(custom_range_c[1])]
        except ValueError:

            raw_input(oops + "CC" + range_exep)
            sys.exit(1)

    if plot_std_toggle_c == 1:

        try:
            plot_std_c = float(plot_std_c)

        except ValueError:

            raw_input(oops + "CC" + plot_exep)
            sys.exit(1)

    if range_toggle_i == 0:

        custom_range_i = custom_range_i.split(",")

        try:
            custom_range_i = [
                           float(custom_range_i[0]),
                           float(custom_range_i[1])]

        except ValueError:

            raw_input(oops + "IRT" + range_exep)
            sys.exit(1)

    if plot_std_toggle_i == 1:

        try:
            plot_std_i = float(plot_std_i)

        except ValueError:

            raw_input(oops + "IRT" + plot_exep)
            sys.exit(1)

    if range_toggle_h == 0:

        custom_range_h = custom_range_h.split(",")

        try:
            custom_range_h = [
                           float(custom_range_h[0]),
                           float(custom_range_h[1])]

        except ValueError:

            raw_input(oops + "Honey" + range_exep)
            sys.exit(1)

    if plot_std_toggle_h == 1:

        try:
            plot_std_h = float(plot_std_h)

        except ValueError:

            raw_input(oops + "Honey" + plot_exep)
            sys.exit(1)

    if range_toggle_p == 0:

        custom_range_p = custom_range_p.split(",")

        try:
            custom_range_p = [
                           float(custom_range_p[0]),
                           float(custom_range_p[1])]

        except ValueError:

            raw_input(oops + "Pep" + range_exep)
            sys.exit(1)

    if plot_std_toggle_p == 1:

        try:
            plot_std_p = float(plot_std_p)

        except ValueError:

            raw_input(oops + "Pep" + plot_exep)
            sys.exit(1)

    # Verify that all the required files exist and are spell correctly
    File_Check(match_folders, cc_toggle, irt_toggle, honey_toggle, pep_toggle)

    for folder_contents in match_folders:

        housing_folder = folder_contents + "\Cleaned_Data\\"

        if os.path.exists(housing_folder) is False:
            os.mkdir(housing_folder)

        Parameters_Executed(housing_folder, inputs)

        # Define names and paths for
        cc_plots = "\CC_plots.csv"
        irt_forward = "\IRT_forward_plots.csv"
        irt_nadir = "\IRT_nadir_plots.csv"
        pep_plots = "\Pep_plots.csv"
        honey_plots = "\Honey_plots.csv"

        cc_plots_txt = "\CC_plots.txt"
        irt_forward_txt = "\IRT_forward_plots.txt"
        irt_nadir_txt = "\IRT_nadir_plots.txt"
        pep_plots_txt = "\Pep_plots.txt"
        honey_plots_txt = "\Honey_plots.txt"

        cc_clean = "\CC_plots_cleaned.csv"
        irt_clean = "\IRT_cleaned.csv"
        honey_clean = "\Honey_plots_cleaned.csv"
        pep_clean = "\Pep_plots_cleaned.csv"

        cc_preen = "CC_plots_preened.csv"
        irt_preen = "IRT_preened.csv"
        pep_preen = "Pep_plots_preened.csv"
        honey_preen = "Honey_plots_preened.csv"

        cc_plot_file = folder_contents + cc_plots
        irt_forward_file = folder_contents + irt_forward
        irt_nadir_file = folder_contents + irt_nadir
        pep_plots_file = folder_contents + pep_plots
        honey_plots_file = folder_contents + honey_plots

        cc_plot_file_txt = folder_contents + cc_plots_txt
        irt_forward_file_txt = folder_contents + irt_forward_txt
        irt_nadir_file_txt = folder_contents + irt_nadir_txt
        pep_plots_file_txt = folder_contents + pep_plots_txt
        honey_plots_file_txt = folder_contents + honey_plots_txt

        clean_cc_csv = folder_contents + cc_clean
        clean_irt_csv = folder_contents + irt_clean
        clean_honey_csv = folder_contents + honey_clean
        clean_pep_csv = folder_contents + pep_clean

        cc_preen_file = housing_folder + cc_preen
        irt_preen_file = housing_folder + irt_preen
        pep_preen_file = housing_folder + pep_preen
        honey_preen_file = housing_folder + honey_preen

        if cc_toggle == 1:
            # Verify the file really exists
            if (os.path.exists(cc_plot_file) or
               os.path.exists(cc_plot_file_txt)) is False:
                raw_input("Missing File! CC Plots file missing from" +
                          housing_folder +
                          "\nBe sure to there is a "
                          "CC_plots file available or "
                          "Do not run the CC_plots section"
                          "in the GUI.\n")
                sys.exit(1)

            # if the files are txt files then change the variable to .txt
            if os.path.exists(cc_plot_file_txt):
                cc_plot_file = cc_plot_file_txt

            # If no reanalyzed toggle then proceed as usual
            if reanalyze_toggle_c == 0:

                Clean_CC_Plot(clean_cc_csv, cc_plot_file,
                              zeros_toggle_c, custom_range_c,
                              range_toggle_c)

            # If reanalyze toggle, then preen the plot file
            elif reanalyze_toggle_c == 1:
                if os.path.exists(clean_cc_csv) is False:
                    Clean_CC_Plot(clean_cc_csv, cc_plot_file,
                                  zeros_toggle_c, custom_range_c,
                                  range_toggle_c)

                Preen_CC_Plot(clean_cc_csv, cc_preen_file)
                Clean_CC_Plot(clean_cc_csv, clean_cc_csv,
                              zeros_toggle_c, custom_range_c,
                              range_toggle_c)

            # Calc the standard deviation and mean
            Stat_Analysis(housing_folder, clean_cc_csv,
                          zeros_toggle_c, reanalyze_toggle_c,
                          plot_std_toggle_c, plot_std_c,
                          range_toggle_c, custom_range_c,
                          stressors, species, graph_toggle_c)

            # execute preening after the stat analysis has been run
            if plot_std_toggle_c == 1:
                Preen_CC_Plot(clean_cc_csv, cc_preen_file)
                Clean_CC_Plot(clean_cc_csv, clean_cc_csv, zeros_toggle_c,
                              custom_range_c, range_toggle_c)

        if irt_toggle == 1:

            if (os.path.exists(irt_forward_file) or
               os.path.exists(irt_forward_file_txt)) is False:
                raw_input("Missing File! IRT Forward file missing from " +
                          housing_folder +
                          "\nBe sure to there is a "
                          "IRT_forward file available or "
                          "Do not run the IRT section in the GUI.\n")
                sys.exit(1)

            if (os.path.exists(irt_nadir_file) or
               os.path.exists(irt_nadir_file_txt)) is False:
                raw_input("Missing File! IRT Nadir file missing from " +
                           housing_folder +
                           "\nBe sure to there is a "
                           "IRT_nadir file available or "
                           "Do not run the IRT section in the GUI.\n")
                sys.exit(1)

            if os.path.exists(irt_forward_file_txt):
                irt_forward_file = irt_forward_file_txt

            if os.path.exists(irt_nadir_file_txt):
                irt_nadir_file = irt_nadir_file_txt

            if reanalyze_toggle_i == 0:

                Clean_IRT_Plot(clean_irt_csv, irt_forward_file,
                               irt_nadir_file, zeros_toggle_i,
                               custom_range_i, range_toggle_i)

            elif reanalyze_toggle_i == 1:
                if os.path.exists(clean_irt_csv) is False:
                    Clean_IRT_Plot(clean_irt_csv, irt_forward_file,
                                   irt_nadir_file, zeros_toggle_i,
                                   custom_range_i, range_toggle_i)

                Preen_CC_Plot(clean_honey_csv, honey_preen_file)
                Clean_IRT_Plot(clean_irt_csv, irt_forward_file,
                               irt_nadir_file, zeros_toggle_i,
                               custom_range_i, range_toggle_i)

            Stat_Analysis(housing_folder, clean_irt_csv,
                          zeros_toggle_i, reanalyze_toggle_i,
                          plot_std_toggle_i, plot_std_i,
                          range_toggle_i, custom_range_i,
                          stressors, species, graph_toggle_i)

            if plot_std_toggle_i == 1:
                Preen_CC_Plot(clean_irt_csv, irt_preen_file)
                Clean_IRT_Plot(clean_irt_csv, clean_irt_csv,
                               clean_irt_csv, zeros_toggle_i,
                               custom_range_i, range_toggle_i)

        if honey_toggle == 1:

            if (os.path.exists(honey_plots_file) or
               os.path.exists(honey_plots_file_txt)) is False:
                raw_input("Missing File!\nHoney Plot file missing from " +
                          housing_folder +
                          "\nBe sure to there is a "
                          "Honey_plots file available or "
                          "Do not run the Honey section in the GUI.\n")
                sys.exit(1)

            if os.path.exists(honey_plots_file_txt):
                honey_plots_file = honey_plots_file_txt

            if reanalyze_toggle_h == 0:

                Clean_HonPep_Plot(clean_honey_csv, honey_plots_file,
                                  zeros_toggle_h, custom_range_h,
                                  range_toggle_h)

            elif reanalyze_toggle_h == 1:
                if os.path.exists(clean_honey_csv) is False:
                    Clean_HonPep_Plot(clean_honey_csv, honey_plots_file,
                                      zeros_toggle_h, custom_range_h,
                                      range_toggle_h)

                Preen_CC_Plot(clean_honey_csv, honey_preen_file)
                Clean_HonPep_Plot(clean_honey_csv, clean_honey_csv,
                                  zeros_toggle_h, custom_range_h,
                                  range_toggle_h)

            Stat_Analysis(housing_folder, clean_honey_csv,
                          zeros_toggle_h, reanalyze_toggle_h,
                          plot_std_toggle_h, plot_std_h,
                          range_toggle_h, custom_range_h,
                          stressors, species, graph_toggle_h)

            if plot_std_toggle_h == 1:
                Preen_CC_Plot(clean_honey_csv, honey_preen_file)
                Clean_HonPep_Plot(clean_honey_csv, clean_honey_csv,
                                  zeros_toggle_h, custom_range_h,
                                  range_toggle_h)

        if pep_toggle == 1:

            if (os.path.exists(pep_plots_file) or
               os.path.exists(pep_plots_file_txt)) is False:
                raw_input("Missing File!\nPep Plot file missing from " +
                           housing_folder +
                           "\nBe sure to there is a "
                           "Pep_plots file available or "
                           "Do not run the Pep section in the GUI.\n")
                sys.exit(1)

            if os.path.exists(pep_plots_file_txt):
                pep_plots_file = pep_plots_file_txt

            if reanalyze_toggle_p == 0:

                Clean_HonPep_Plot(clean_pep_csv, pep_plots_file,
                                  zeros_toggle_p, custom_range_p,
                                  range_toggle_p)

            elif reanalyze_toggle_p == 1:
                if os.path.exists(clean_pep_csv) is False:
                    Clean_HonPep_Plot(clean_pep_csv, pep_plots_file,
                                      zeros_toggle_p, custom_range_p,
                                      range_toggle_p)

                Preen_CC_Plot(clean_pep_csv, pep_preen_file)
                Clean_HonPep_Plot(clean_pep_csv, clean_pep_csv,
                                  zeros_toggle_p, custom_range_p,
                                  range_toggle_p)

            Stat_Analysis(housing_folder, clean_pep_csv,
                          zeros_toggle_p, reanalyze_toggle_p,
                          plot_std_toggle_p, plot_std_p,
                          range_toggle_p, custom_range_p,
                          stressors, species, graph_toggle_p)

            if plot_std_toggle_h == 1:
                Preen_CC_Plot(clean_pep_csv, pep_preen_file)
                Clean_HonPep_Plot(clean_pep_csv, clean_pep_csv, zeros_toggle_p,
                                  custom_range_p, range_toggle_p)

    #Combine_Stats(top_folder)
    #Clean_Up(top_folder)

if __name__ == '__main__':
    main()
